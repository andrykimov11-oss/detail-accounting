"""
Импорт операторов (физлиц) из выгрузки 1С в справочник системы.

Операторы входят в систему по имени без пароля (docs/roadmap.md), поэтому
справочник должен быть заполнен до пилота — иначе оператору некем войти.
Источник — колонка «Исполнитель» в выгрузке «производственные операции».

## Что делает

1. Читает уникальных исполнителей из выгрузки 1С.
2. Генерирует стабильный operator_id (транслитерация ФИО).
3. Заносит в таблицу operators.
4. **Сомнительные записи не тащит молча**, а показывает отдельно на ревью:
   служебные значения («asv»), слишком короткие, без заглавной буквы.

## Фильтр под пилот

Опция `--area-operations` ограничивает импорт операторами, которые реально
работали на операциях заданного участка. Для пилота на кромлении это заведёт
не всех 39 человек, а только кромлёнщиков — короче список на экране входа.

Применение:
    python src/import_operators.py --db prod.db \\
        --plan "производственные операции.xlsx"

    python src/import_operators.py --db prod.db \\
        --plan "производственные операции.xlsx" \\
        --areas "Операции по участкам.xlsx" --area Кромление
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from storage import Storage  # noqa: E402

# Транслитерация кириллицы в латиницу для читаемого operator_id
TRANSLIT = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e",
    "ж": "zh", "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m",
    "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
    "ф": "f", "х": "h", "ц": "c", "ч": "ch", "ш": "sh", "щ": "sch",
    "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
}


@dataclass
class ImportReport:
    """Итог импорта операторов."""
    imported: list[tuple[str, str]] = field(default_factory=list)  # (id, ФИО)
    review: list[tuple[str, int]] = field(default_factory=list)    # (значение, частота)
    total_source: int = 0
    filtered_by_area: int = 0


def transliterate(text: str) -> str:
    out = []
    for ch in text.lower():
        out.append(TRANSLIT.get(ch, ch if ch.isalnum() else " "))
    return "".join(out)


def make_operator_id(full_name: str, taken: set[str]) -> str:
    """
    Стабильный читаемый ID: op_<фамилия>_<инициалы>.
    При коллизии добавляется счётчик — двух одинаковых ID не будет.

        «Мазеин Петр Иванович» → op_mazein_p_i
        «Малышев Андрей Юрьевич» → op_malyshev_a_yu
    """
    words = [w for w in transliterate(full_name).split() if w]
    if not words:
        base = "op_" + re.sub(r"[^a-z0-9]", "", transliterate(full_name))[:8]
    else:
        surname = words[0]
        initials = "_".join(w[0] for w in words[1:3])
        base = "op_" + surname + (("_" + initials) if initials else "")
    base = re.sub(r"_+", "_", base).strip("_")

    candidate = base
    n = 2
    while candidate in taken:
        candidate = f"{base}_{n}"
        n += 1
    taken.add(candidate)
    return candidate


def looks_like_name(value: str) -> bool:
    """
    Похоже ли значение на ФИО/бригаду, а не на служебный мусор.

    Настоящее имя: есть слово с заглавной буквы длиной от 3 символов.
    Отсеиваются «asv», пустые, чисто цифровые.
    «Бригада (Саня Сергей)» проходит — это легитимная рабочая единица.
    """
    s = (value or "").strip()
    if len(s) < 3:
        return False
    words = re.findall(r"[A-Za-zА-Яа-яЁё]{3,}", s)
    return any(w[0].isupper() for w in words)


def read_executors(plan_path: Path) -> Counter:
    """Уникальные исполнители из выгрузки 1С с частотой."""
    import openpyxl

    wb = openpyxl.load_workbook(plan_path, data_only=True, read_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(h or "").strip().lower() for h in next(rows)]
        idx = next((i for i, h in enumerate(header) if "исполнитель" in h), None)
        if idx is None:
            raise ValueError("в выгрузке нет колонки «Исполнитель»")

        execs: Counter = Counter()
        for row in rows:
            if idx < len(row) and row[idx]:
                execs[str(row[idx]).strip()] += 1
        return execs
    finally:
        wb.close()


def executors_of_area(plan_path: Path, area_map: dict[str, str],
                      area: str) -> set[str]:
    """
    Исполнители, работавшие на операциях заданного участка.
    Нужно, чтобы завести под пилот только операторов этого участка.
    """
    import openpyxl

    ops_of_area = {op for op, a in area_map.items() if a == area}
    wb = openpyxl.load_workbook(plan_path, data_only=True, read_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(h or "").strip().lower() for h in next(rows)]
        ex_idx = next((i for i, h in enumerate(header) if "исполнитель" in h), None)
        op_idx = next((i for i, h in enumerate(header) if "операция цеха" in h), None)
        if ex_idx is None or op_idx is None:
            return set()

        result: set[str] = set()
        for row in rows:
            if op_idx < len(row) and ex_idx < len(row):
                op = str(row[op_idx] or "").strip()
                ex = str(row[ex_idx] or "").strip()
                if op in ops_of_area and ex:
                    result.add(ex)
        return result
    finally:
        wb.close()


def import_operators(storage: Storage, plan_path: Path,
                     area: str | None = None,
                     areas_path: Path | None = None) -> ImportReport:
    """
    Прочитать исполнителей и завести в справочник операторов.

    area + areas_path — необязательный фильтр: только операторы участка.
    """
    report = ImportReport()
    execs = read_executors(plan_path)
    report.total_source = len(execs)

    allow: set[str] | None = None
    if area and areas_path:
        from one_c_loader import load_area_operations  # noqa: PLC0415
        area_map = load_area_operations(areas_path)
        allow = executors_of_area(plan_path, area_map, area)
        report.filtered_by_area = len(allow)

    taken: set[str] = set()
    for name, count in execs.most_common():
        if allow is not None and name not in allow:
            continue
        if not looks_like_name(name):
            report.review.append((name, count))
            continue
        operator_id = make_operator_id(name, taken)
        storage.upsert_operator(operator_id, name, is_active=True)
        report.imported.append((operator_id, name))

    return report


def format_report(report: ImportReport, area: str | None) -> str:
    lines = ["=== ИМПОРТ ОПЕРАТОРОВ ===", ""]
    lines.append(f"Исполнителей в выгрузке: {report.total_source}")
    if area:
        lines.append(f"Работали на участке «{area}»: {report.filtered_by_area}")
    lines.append(f"Заведено операторов: {len(report.imported)}")
    lines.append("")
    for oid, name in report.imported:
        lines.append(f"  {oid:<22} {name}")
    if report.review:
        lines.append("")
        lines.append(f"⚠ НА РЕВЬЮ (не похоже на ФИО, не заведены): {len(report.review)}")
        for value, count in report.review:
            lines.append(f"  «{value}»  ({count} записей)")
        lines.append("  Если это реальные люди — завести вручную "
                     "(storage.upsert_operator).")
    return "\n".join(lines)


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Импорт операторов из выгрузки 1С")
    p.add_argument("--db", required=True, help="файл БД")
    p.add_argument("--plan", required=True, help="xlsx «производственные операции»")
    p.add_argument("--areas", default=None, help="xlsx «Операции по участкам» (для --area)")
    p.add_argument("--area", default=None, help="завести только операторов этого участка")
    args = p.parse_args(argv)

    if args.area and not args.areas:
        print("--area требует --areas (справочник операций по участкам)",
              file=sys.stderr)
        return 1

    storage = Storage(args.db)
    try:
        report = import_operators(
            storage, Path(args.plan),
            area=args.area,
            areas_path=Path(args.areas) if args.areas else None,
        )
        print(format_report(report, args.area))
        return 0
    finally:
        storage.close()


if __name__ == "__main__":
    sys.exit(main())
