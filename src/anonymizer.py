"""
Обезличиватель производственных данных (ФЗ-152).

Назначение: убрать персональные данные (ФИО клиентов, телефоны, названия
компаний) из выгрузок 1С, .xbir и папок заказов, чтобы данные можно было
безопасно передавать для анализа/тестирования.

Принципы:
1. Детерминированность. Одно и то же имя клиента → один и тот же псевдоним
   во всех файлах (чтобы связи сохранялись). Достигается через общий словарь.
2. Сохранение структуры. Формат файла не меняется, заменяются только
   чувствительные значения. .xlsx остаётся .xlsx, .xbir остаётся .xbir.
3. Аудит. Словарь замен (клиент → псевдоним) сохраняется в отдельный файл
   anonymization_map.json — он нужен, чтобы при необходимости связать
   псевдоним обратно с реальным клиентом (для разработчика, не для публики).

Применение:
    python src/anonymizer.py <входная_папка> <выходная_папка>

    Обрабатывает:
    - *.xlsx (отчёты 1С): убирает ФИО, телефоны
    - *.xbir (Базис-раскрой): убирает имя клиента из "Номер заказа"
    - имена папок/файлов: убирает название клиента из пути
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import sys
import xml.etree.ElementTree as ET
from pathlib import Path


class Anonymizer:
    """
    Обезличиватель. Хранит словарь замен и применяет его ко всем файлам.

    Словарь pseudonyms: {реальное_имя_клиента → псевдоним}.
    Между запусками словарь сохраняется в map_path (JSON).
    """

    # Шаблоны извлечения имён клиентов из разных источников
    PHONE_RE = re.compile(r"\+?\d?[\s\-]?\(?\d{3}\)?[\s\-]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}")

    def __init__(self, map_path: Path | None = None):
        self.map_path = map_path
        self.pseudonyms: dict[str, str] = {}
        if map_path and map_path.exists():
            self.pseudonyms = json.loads(map_path.read_text(encoding="utf-8"))

    def _save_map(self) -> None:
        if self.map_path:
            self.map_path.write_text(
                json.dumps(self.pseudonyms, ensure_ascii=False, indent=2),
                encoding="utf-8"
            )

    def _pseudonym_for(self, real_name: str) -> str:
        """Вернуть псевдоним для имени клиента (детерминированно)."""
        if not real_name or not real_name.strip():
            return real_name
        key = real_name.strip()
        if key not in self.pseudonyms:
            # Стабильный номер на основе хеша, чтобы при повторных запусках
            # получать тот же номер для того же имени
            num = int(hashlib.md5(key.encode()).hexdigest(), 16) % 10000
            self.pseudonyms[key] = f"Клиент_{num:04d}"
            self._save_map()
        return self.pseudonyms[key]

    # --- Обезличивание строк ------------------------------------------------

    def _clean_phone(self, text: str) -> str:
        """Заменить телефоны на [PHONE]."""
        return self.PHONE_RE.sub("[PHONE]", text or "")

    def _clean_client_name(self, text: str) -> str:
        """Заменить имя клиента на псевдоним (если строка похожа на ФИО/ООО)."""
        if not text or not text.strip():
            return text
        # Числовые/служебные значения не трогаем
        s = text.strip()
        if re.fullmatch(r"\d+[\.,]?\d*", s):
            return text
        # Если это "6564 Спецторг ООО" — заменим только часть с именем
        return self._pseudonym_for(s)

    # --- Обработка .xlsx (отчёт 1С) -----------------------------------------

    def process_xlsx(self, src: Path, dst: Path) -> dict:
        """Обезличить xlsx-отчёт 1С. Возвращает статистику замен."""
        import openpyxl
        wb = openpyxl.load_workbook(src, data_only=True)
        stats = {"phones": 0, "clients": 0, "cells": 0}

        # Колонки с ПД (ищем по подстроке в заголовке)
        client_headers = ["клиент", "заказ клиента", "покупател"]
        phone_headers = ["телефон", "тел."]

        for ws in wb.worksheets:
            headers = [str(c.value).lower() if c.value else "" for c in ws[1]]

            for col_idx, header in enumerate(headers):
                is_client = any(h in header for h in client_headers)
                is_phone = any(h in header for h in phone_headers)
                if not (is_client or is_phone):
                    continue

                for row in ws.iter_rows(min_row=2, min_col=col_idx + 1,
                                        max_col=col_idx + 1):
                    for cell in row:
                        if cell.value is None:
                            continue
                        original = str(cell.value)
                        if is_phone:
                            cleaned = self._clean_phone(original)
                            if cleaned != original:
                                stats["phones"] += 1
                        else:
                            cleaned = self._clean_client_name(original)
                            if cleaned != original:
                                stats["clients"] += 1
                        cell.value = cleaned
                        stats["cells"] += 1

        dst.parent.mkdir(parents=True, exist_ok=True)
        wb.save(dst)
        return stats

    # --- Обработка .xbir (Базис-раскрой) ------------------------------------

    def process_xbir(self, src: Path, dst: Path) -> dict:
        """Обезличить .xbir: убрать имя клиента из 'Номер заказа'."""
        stats = {"clients": 0, "rows": 0}
        raw = src.read_bytes()
        if raw.startswith(b"\xef\xbb\xbf"):
            raw = raw[3:]
        root = ET.fromstring(raw)

        # Найти колонку "Номер заказа" (в .xbir значение содержит "6564 Спецторг ООО")
        cols = root.findall(".//Col")
        name_to_pos = {c.get("Name"): int(c.get("Index", "0")) - 1 for c in cols}
        order_pos = name_to_pos.get("Номер заказа")
        if order_pos is None:
            # нет колонки — нечего обезличивать
            dst.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src, dst)
            return stats

        for row_el in root.findall(".//Row"):
            text = row_el.text or ""
            cells = text.split("\t")
            if order_pos < len(cells):
                original = cells[order_pos]
                # "6564 Спецторг ООО" → извлечь имя после первого числа
                parts = original.split(None, 1)
                if len(parts) >= 2:
                    num_part = parts[0]
                    name_part = parts[1]
                    pseudo = self._pseudonym_for(name_part)
                    cells[order_pos] = f"{num_part} {pseudo}"
                    if pseudo != name_part:
                        stats["clients"] += 1
                    row_el.text = "\t".join(cells)
                    stats["rows"] += 1

        dst.parent.mkdir(parents=True, exist_ok=True)
        # Сохраним с BOM, как в оригинале
        xml_bytes = b"\xef\xbb\xbf" + ET.tostring(root, encoding="UTF-8")
        dst.write_bytes(xml_bytes)
        return stats

    # --- Обработка имени папки/файла ----------------------------------------

    def clean_path_component(self, component: str) -> str:
        """Заменить название клиента в имени папки/файла."""
        # "6564-Spectorg-OOO" → "6564-Клиент_1234"
        # Шаблон: <число>-<текст-с-дефисами>
        m = re.match(r"^(\d+)-(.+)$", component)
        if m:
            num, name = m.group(1), m.group(2)
            # Уберём лишние дефисы для сопоставления, восстановим псевдоним
            pseudo = self._pseudonym_for(name.replace("-", " "))
            return f"{num}-{pseudo}"
        return component

    # --- Рекурсивная обработка папки ----------------------------------------

    def process_directory(self, src_dir: Path, dst_dir: Path) -> dict:
        """Обезличить всю папку рекурсивно."""
        total = {"files": 0, "phones": 0, "clients": 0, "cells": 0, "rows": 0}

        for item in src_dir.iterdir():
            # Пропускаем системный мусор, НО сохраняем .xbir (файлы Базис-раскроя)
            # — у них имя начинается с точки, но это полноценные данные.
            if item.name in (".DS_Store", ".git", ".gitignore"):
                continue
            # Вычисляем безопасное имя
            clean_name = self.clean_path_component(item.name)
            dst_item = dst_dir / clean_name

            if item.is_dir():
                sub = self.process_directory(item, dst_item)
                for k, v in sub.items():
                    total[k] = total.get(k, 0) + v
            elif item.is_file():
                if item.suffix.lower() == ".xlsx":
                    s = self.process_xlsx(item, dst_item)
                    total["phones"] += s["phones"]
                    total["clients"] += s["clients"]
                    total["cells"] += s["cells"]
                elif item.suffix.lower() == ".xbir":
                    s = self.process_xbir(item, dst_item)
                    total["clients"] += s["clients"]
                    total["rows"] += s["rows"]
                else:
                    # прочие файлы (bmp, xPrg) копируем как есть
                    dst_item.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(item, dst_item)
                total["files"] += 1

        return total


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Обезличиватель данных (ФЗ-152)")
    p.add_argument("src", help="входная папка с реальными данными")
    p.add_argument("dst", help="выходная папка для обезличенных данных")
    p.add_argument("--map", dest="map_path", default=None,
                   help="путь к JSON-словарю замен (для переиспользования)")
    args = p.parse_args(argv)

    src = Path(args.src)
    dst = Path(args.dst)
    if not src.is_dir():
        print(f"Ошибка: входная папка не найдена: {src}", file=sys.stderr)
        return 1

    map_path = Path(args.map_path) if args.map_path else dst / "anonymization_map.json"
    anon = Anonymizer(map_path)

    print(f"Обезличивание: {src} → {dst}")
    print(f"Словарь замен: {map_path}")
    print()
    stats = anon.process_directory(src, dst)

    print(f"Готово:")
    print(f"  файлов обработано: {stats['files']}")
    print(f"  телефонов скрыто:  {stats['phones']}")
    print(f"  имён клиентов:     {stats['clients']}")
    print(f"  ячеек изменено:    {stats['cells']}")
    print(f"  строк в .xbir:     {stats['rows']}")
    print(f"  псевдонимов в словаре: {len(anon.pseudonyms)}")
    print()
    print(f"Обезличенные данные: {dst}")
    print(f"Словарь (только для разработчика!): {map_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
