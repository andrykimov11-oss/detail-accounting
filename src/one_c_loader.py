"""
Загрузчик выгрузок 1С.

Читает выгрузку «производственные операции» и превращает её в модель заказа
1С: полный номер документа, клиент, статус, срок, список операций цеха.

Ключевая деталь — **номер заказа в 1С состоит из префикса и числа**
(`ЛД00-006564`), а в `.xbir` из Базиса приходит только число (`6564`).
Префикс при переносе в Базис теряется. Загрузчик разбирает номер на части,
чтобы резолвер (`order_resolver.py`) мог восстановить связку.

Проверено на массиве: 3 префикса (`ПС00`, `ЛД00`, `0Ч00`), 10 586 заказов,
102 878 строк операций.

Применение:
    orders = load_production_plan(Path("производственные операции.xlsx"))
    orders[6564]  # → [OneCOrder(...), ...] — кандидаты с этим числом
"""
from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Optional

# Номер документа 1С: префикс + дефис + число («ЛД00-006564», «0Ч00-001153»)
ORDER_NUM_RE = re.compile(r"^(?P<prefix>.+?)-(?P<digits>\d+)$")

# Имена колонок выгрузки. Ищутся по подстроке, регистр не важен —
# состав колонок между выгрузками 1С слегка плавает.
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "order_full_num": ("номер заказа",),
    "client_name": ("клиент", "наименование клиента"),
    "order_status": ("заказ клиента.статус", "статус заказа"),
    "operation_name": ("операция цеха", "операция"),
    "order_date": ("дата заказа",),
    "deadline": ("плановая дата выдачи",),
    "manager": ("менеджер",),
    "executor": ("исполнитель",),
}

# Статусы заказов, которые не идут в производство
CLOSED_STATUSES = ("закрыт",)


@dataclass
class OneCOrder:
    """Заказ 1С: документ с номером, клиентом и списком операций цеха."""
    order_full_num: str                  # 'ЛД00-006564' — полный номер документа
    order_prefix: str                    # 'ЛД00'
    order_num: int                       # 6564 — числовая часть (связь с Базисом)
    client_name: str = ""
    order_status: str = ""
    order_date: Optional[date] = None
    deadline: Optional[date] = None
    manager: str = ""
    operations: list[str] = field(default_factory=list)

    @property
    def is_active(self) -> bool:
        """Заказ в работе (не закрыт) — только такие нужны цеху."""
        return not any(s in self.order_status.lower() for s in CLOSED_STATUSES)


@dataclass
class LoadReport:
    """Итог чтения выгрузки."""
    source_file: str = ""
    rows_total: int = 0
    rows_skipped: int = 0
    orders_total: int = 0
    prefixes: dict[str, int] = field(default_factory=dict)
    operations_total: int = 0
    unparsed_numbers: list[str] = field(default_factory=list)
    missing_columns: list[str] = field(default_factory=list)


def parse_order_number(raw: str) -> tuple[str, int] | None:
    """
    Разобрать номер документа 1С на префикс и число.

        'ЛД00-006564'  → ('ЛД00', 6564)
        'ПС00-007730'  → ('ПС00', 7730)
        '6564'         → ('', 6564)          — номер без префикса
        'мусор'        → None

    Ведущие нули числовой части срезаются: именно в таком виде номер
    попадает в Базис, и по нему идёт сопоставление.
    """
    s = (raw or "").strip()
    if not s:
        return None

    m = ORDER_NUM_RE.match(s)
    if m:
        return m.group("prefix").strip(), int(m.group("digits"))

    if s.isdigit():
        return "", int(s)

    return None


def _resolve_columns(header: Iterable) -> tuple[dict[str, int], list[str]]:
    """
    Сопоставить колонки выгрузки с полями модели.

    Порядок проверки важен: сначала **точное** совпадение имени колонки,
    и только потом вхождение подстроки. Иначе поле «Клиент» подхватывает
    колонку «Заказ клиента» (там лежит строка вида «Заказ клиента
    ЛД00-007936 от 30.07.2024»), и резолвер сравнивает фамилию с описанием
    документа — связка не находится никогда.

    Псевдонимы внутри поля тоже перебираются по приоритету: «Плановая дата
    выдачи» важнее, чем «Срок изготовления».
    """
    cells = [str(h or "").strip().lower() for h in header]
    index: dict[str, int] = {}

    for field_name, aliases in COLUMN_ALIASES.items():
        for alias in aliases:                      # приоритет псевдонима
            exact = next((i for i, c in enumerate(cells) if c == alias), None)
            if exact is not None:
                index[field_name] = exact
                break
        else:
            for alias in aliases:                  # затем — вхождение
                partial = next((i for i, c in enumerate(cells) if alias in c), None)
                if partial is not None:
                    index[field_name] = partial
                    break

    missing = [f for f in ("order_full_num", "client_name") if f not in index]
    return index, missing


def _as_date(value) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%Y %H:%M:%S"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def load_production_plan(
    path: Path,
    active_only: bool = True,
) -> tuple[dict[int, list[OneCOrder]], LoadReport]:
    """
    Прочитать выгрузку «производственные операции» из 1С.

    Возвращает:
        (orders_by_num, report)

        orders_by_num — числовая часть номера → список заказов 1С с этим
        числом. Список длиннее одного означает неоднозначность: в Базисе
        такое число соответствует нескольким документам 1С, и связку
        должен разрешить резолвер.

    Одна строка выгрузки = одна операция заказа, поэтому заказы
    собираются по полному номеру, а операции накапливаются.
    """
    import openpyxl

    report = LoadReport(source_file=str(path))
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)

        try:
            header = next(rows)
        except StopIteration:
            return {}, report

        idx, missing = _resolve_columns(header)
        report.missing_columns = missing
        if missing:
            raise ValueError(
                f"в выгрузке нет обязательных колонок: {', '.join(missing)}"
            )

        def cell(row, key):
            i = idx.get(key)
            return row[i] if i is not None and i < len(row) else None

        # ключ: (полный номер, дата заказа) — см. комментарий ниже
        by_full: dict[tuple[str, str], OneCOrder] = {}

        for row in rows:
            report.rows_total += 1

            raw_num = cell(row, "order_full_num")
            parsed = parse_order_number(str(raw_num or ""))
            if parsed is None:
                report.rows_skipped += 1
                if raw_num and len(report.unparsed_numbers) < 50:
                    report.unparsed_numbers.append(str(raw_num))
                continue

            prefix, num = parsed
            full = str(raw_num).strip()
            row_date = _as_date(cell(row, "order_date"))

            # Ключ заказа — пара «полный номер + дата», а НЕ номер сам по себе.
            # Номера документов в 1С переиспользуются по годам: 0Ч00-001991
            # встречается в 2024, 2025 и 2026 годах у трёх разных клиентов,
            # и все три записи активны. Группировка по одному номеру
            # склеивала бы разные заказы в один и подставляла чужого клиента.
            key = (full, row_date.isoformat() if row_date else "")

            order = by_full.get(key)
            if order is None:
                order = OneCOrder(
                    order_full_num=full,
                    order_prefix=prefix,
                    order_num=num,
                    client_name=str(cell(row, "client_name") or "").strip(),
                    order_status=str(cell(row, "order_status") or "").strip(),
                    order_date=row_date,
                    deadline=_as_date(cell(row, "deadline")),
                    manager=str(cell(row, "manager") or "").strip(),
                )
                by_full[key] = order

            operation = cell(row, "operation_name")
            if operation:
                op = str(operation).strip()
                if op not in order.operations:
                    order.operations.append(op)
                    report.operations_total += 1
    finally:
        wb.close()

    orders = list(by_full.values())
    if active_only:
        orders = [o for o in orders if o.is_active]

    by_num: dict[int, list[OneCOrder]] = defaultdict(list)
    for o in orders:
        by_num[o.order_num].append(o)
        report.prefixes[o.order_prefix] = report.prefixes.get(o.order_prefix, 0) + 1

    report.orders_total = len(orders)
    return dict(by_num), report


def load_area_operations(path: Path) -> dict[str, str]:
    """
    Прочитать справочник «Операция цеха → Участок» из выгрузки 1С.

    В цехе 227 операций и 6 участков — этот справочник заменяет
    захардкоженные правила (см. docs/dataset-validation-report.md, H3).
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        mapping: dict[str, str] = {}
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            try:
                header = [str(h or "").strip().lower() for h in next(rows)]
            except StopIteration:
                continue

            op_idx = next((i for i, h in enumerate(header) if "операция" in h), None)
            area_idx = next((i for i, h in enumerate(header) if "участок" in h), None)
            if op_idx is None or area_idx is None:
                continue

            for row in rows:
                if op_idx >= len(row) or area_idx >= len(row):
                    continue
                op, area = row[op_idx], row[area_idx]
                if op and area:
                    mapping[str(op).strip()] = str(area).strip()
        return mapping
    finally:
        wb.close()
