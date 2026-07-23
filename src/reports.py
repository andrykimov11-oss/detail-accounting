"""
Отчёты по итогам работы — для проверки отработки на пилоте (Фаза E).

Пилот проверяет, правильно ли система считает: какой заказ прошёл, сколько
деталей, кто сканировал, где аномалии. Эти отчёты дают факты для решения,
идти ли в Фазу F.

Три отчёта:
  order_report   — по одному заказу: операции, план/факт, статус, потери;
  shift_report   — по смене/дню: сколько заказов, деталей, аномалий, по кому;
  operators_report — по операторам: сколько деталей отсканировал каждый.

Данные берутся из scan_events (audit log) и facts — всё, что накопила БД.
"""
from __future__ import annotations

import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from operation_plan import build_order_plan  # noqa: E402
from status_calc import calc_operation_status  # noqa: E402


# --- Отчёт по заказу ---------------------------------------------------------

@dataclass
class OperationLine:
    operation_1c: str
    planned: int
    scanned: int
    status: str
    lost: int          # сколько деталей не отсканировано

    @property
    def pct(self) -> float:
        return round(100 * self.scanned / self.planned, 1) if self.planned else 0.0


@dataclass
class OrderReport:
    order_num: int
    order_full_num: str = ""
    client_name: str = ""
    operations: list[OperationLine] = field(default_factory=list)

    @property
    def planned_total(self) -> int:
        return sum(o.planned for o in self.operations)

    @property
    def scanned_total(self) -> int:
        return sum(o.scanned for o in self.operations)


def order_report(core, order_num: int, operations_1c: list[str] | None = None,
                 rules: dict | None = None) -> OrderReport:
    """
    Отчёт план/факт по одному заказу.

    Если список операций не передан, он берётся из факта — операций, по
    которым были сканы этого заказа. Так отчёт отвечает на вопрос «что
    реально происходило», не требуя внешнего списка операций.
    """
    ctx = core.load_order(order_num)
    link = core.storage.get_order_link(order_num)
    rep = OrderReport(
        order_num=order_num,
        order_full_num=(link["order_full_num"] if link else ""),
        client_name=(link["client_name"] if link else ""),
    )

    if operations_1c is None:
        facts = core.storage.get_facts_by_order(order_num)
        operations_1c = sorted({f["operation_1c"] for f in facts})

    plans, _ = build_order_plan(ctx.details, operations_1c, rules=rules)
    for plan in plans:
        if not plan.is_per_detail:
            continue
        st = calc_operation_status(plan, ctx)
        rep.operations.append(OperationLine(
            operation_1c=plan.operation_1c,
            planned=st.planned_total,
            scanned=st.scanned_total,
            status=st.status.value,
            lost=len(st.lost_uids),
        ))
    return rep


# --- Отчёт по смене/дню ------------------------------------------------------

@dataclass
class ShiftReport:
    date_from: str = ""
    date_to: str = ""
    orders: set = field(default_factory=set)
    scans_total: int = 0
    accepted: int = 0
    duplicate: int = 0
    overplan: int = 0
    errors: int = 0
    by_operator: Counter = field(default_factory=Counter)     # оператор → принято
    by_operation: Counter = field(default_factory=Counter)    # операция → принято

    @property
    def anomalies(self) -> int:
        """Аномалии: превышение плана + нераспознанные + ошибки."""
        return self.overplan + self.errors

    @property
    def anomaly_pct(self) -> float:
        return round(100 * self.anomalies / self.scans_total, 1) if self.scans_total else 0.0


def shift_report(core, day_from: date | None = None,
                 day_to: date | None = None) -> ShiftReport:
    """
    Сводка по сканам за период (по умолчанию — все).
    Аномалии и распределение по операторам/операциям — метрики пилота.
    """
    rep = ShiftReport(
        date_from=day_from.isoformat() if day_from else "",
        date_to=day_to.isoformat() if day_to else "",
    )
    rows = core.storage._conn.execute(
        "SELECT * FROM scan_events ORDER BY scanned_at"
    ).fetchall()

    for r in rows:
        ts = (r["scanned_at"] or "")[:10]
        if day_from and ts and ts < day_from.isoformat():
            continue
        if day_to and ts and ts > day_to.isoformat():
            continue

        rep.scans_total += 1
        status = r["status"]
        if status == "accepted":
            rep.accepted += 1
            if r["operator_id"]:
                rep.by_operator[r["operator_id"]] += 1
            if r["operation_1c"]:
                rep.by_operation[r["operation_1c"]] += 1
        elif status == "duplicate":
            rep.duplicate += 1
        elif status == "overplan":
            rep.overplan += 1
        else:
            rep.errors += 1

    # заказы периода — по деталям принятых сканов
    for r in rows:
        if r["status"] == "accepted" and r["detail_uid"]:
            drow = core.storage._conn.execute(
                "SELECT order_num FROM details WHERE detail_uid=? LIMIT 1",
                (r["detail_uid"],)).fetchone()
            if drow:
                rep.orders.add(drow["order_num"])
    return rep


# --- Текстовые отчёты --------------------------------------------------------

def format_order_report(rep: OrderReport) -> str:
    lines = [f"=== ОТЧЁТ ПО ЗАКАЗУ {rep.order_num} ===", ""]
    if rep.order_full_num:
        lines.append(f"Документ 1С: {rep.order_full_num}   Клиент: {rep.client_name}")
        lines.append("")
    lines.append(f"{'Операция':<40} {'план':>6} {'факт':>6} {'%':>6}  статус")
    for o in rep.operations:
        lines.append(f"{o.operation_1c[:40]:<40} {o.planned:>6} {o.scanned:>6} "
                     f"{o.pct:>5}%  {o.status}"
                     + (f"  ⚠ потеряно {o.lost}" if o.lost else ""))
    lines.append("")
    lines.append(f"ИТОГО: {rep.scanned_total} из {rep.planned_total} деталей")
    return "\n".join(lines)


def format_shift_report(rep: ShiftReport) -> str:
    lines = ["=== ОТЧЁТ ПО СМЕНЕ ===", ""]
    period = " ".join(x for x in (rep.date_from, "—", rep.date_to) if x.strip("—"))
    if period.strip(" —"):
        lines.append(f"Период: {period}")
    lines.append(f"Заказов затронуто:  {len(rep.orders)}")
    lines.append(f"Сканов всего:       {rep.scans_total}")
    lines.append(f"  принято:          {rep.accepted}")
    lines.append(f"  повторов:         {rep.duplicate}")
    lines.append(f"  превышений:       {rep.overplan}")
    lines.append(f"  ошибок:           {rep.errors}")
    lines.append(f"Аномалий:           {rep.anomalies} ({rep.anomaly_pct}%)")
    lines.append("")
    lines.append("По операторам (принято деталей):")
    for op, n in rep.by_operator.most_common():
        lines.append(f"  {n:>6}  {op}")
    lines.append("")
    lines.append("По операциям (принято деталей):")
    for op, n in rep.by_operation.most_common():
        lines.append(f"  {n:>6}  {op}")
    return "\n".join(lines)


# --- Выгрузка в xlsx (для журнала пилота) ------------------------------------

def export_shift_xlsx(rep: ShiftReport, out_path: Path) -> Path:
    """Выгрузить сводку смены в xlsx — для журнала и согласования."""
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сводка смены"
    bold = Font(bold=True)

    ws["A1"] = "Отчёт по смене"; ws["A1"].font = bold
    ws["A2"] = "Период"; ws["B2"] = f"{rep.date_from} — {rep.date_to}".strip(" —")
    rows = [
        ("Заказов затронуто", len(rep.orders)),
        ("Сканов всего", rep.scans_total),
        ("Принято", rep.accepted),
        ("Повторов", rep.duplicate),
        ("Превышений", rep.overplan),
        ("Ошибок", rep.errors),
        ("Аномалий", rep.anomalies),
        ("Аномалий, %", rep.anomaly_pct),
    ]
    for i, (k, v) in enumerate(rows, start=4):
        ws.cell(i, 1, k); ws.cell(i, 2, v)

    ws2 = wb.create_sheet("По операторам")
    ws2.append(["Оператор", "Принято деталей"])
    ws2["A1"].font = bold; ws2["B1"].font = bold
    for op, n in rep.by_operator.most_common():
        ws2.append([op, n])

    ws3 = wb.create_sheet("По операциям")
    ws3.append(["Операция", "Принято деталей"])
    ws3["A1"].font = bold; ws3["B1"].font = bold
    for op, n in rep.by_operation.most_common():
        ws3.append([op, n])

    for w in (ws, ws2, ws3):
        w.column_dimensions["A"].width = 32
        w.column_dimensions["B"].width = 18

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
