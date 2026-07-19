"""
Валидация гипотез спецификаций на реальном массиве данных (Этап C, gate).

Все спецификации проекта написаны по анализу ОДНОГО заказа (Спецторг,
64 записи, 47 уникальных деталей). Этот скрипт проверяет, выдерживают ли
сделанные из него выводы полный массив цеха.

Проверяет пять гипотез:

  H1. QR-код MD5(GUID)[:10] уникален          → иначе учёт неоднозначен
  H2. Номер заказа извлекается однозначно      → корневой риск проекта
  H3. Правила операций покрывают реальный цех  → иначе план не посчитать
  H4. Распределение qty допускает поштучный скан → иначе нужен режим B
  H5. Данные .xbir пригодны для учёта          → доля битых строк

Применение:
    python src/validate_dataset.py samples/ --out out/
    python src/validate_dataset.py samples/ --out out/ --plan production_plan.xlsx
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from operation_mapping import OPERATION_RULES, find_rule  # noqa: E402
from xbir_parser import Detail, parse_xbir  # noqa: E402

QR_LENGTHS = (8, 10, 12)


@dataclass
class DatasetReport:
    """Итог валидации массива."""
    files_total: int = 0
    files_failed: list[str] = field(default_factory=list)
    rows_total: int = 0
    rows_parsed: int = 0
    row_errors: list[str] = field(default_factory=list)

    orders: dict[int, int] = field(default_factory=dict)       # заказ → деталей
    order_raw_patterns: Counter = field(default_factory=Counter)
    order_num_failures: list[str] = field(default_factory=list)

    qr_collisions: dict[int, list[tuple[str, str, str]]] = field(default_factory=dict)
    uid_total: int = 0

    qty_distribution: Counter = field(default_factory=Counter)
    thickness_distribution: Counter = field(default_factory=Counter)
    edge_distribution: Counter = field(default_factory=Counter)

    operations_in_plan: Counter = field(default_factory=Counter)
    operations_without_rule: list[str] = field(default_factory=list)
    rules_without_details: list[str] = field(default_factory=list)

    details_no_operation: int = 0
    drill_details: int = 0


# --- H1: коллизии QR ---------------------------------------------------------

def check_qr_collisions(uids: set[str]) -> dict[int, list[tuple[str, str, str]]]:
    """
    Проверить уникальность MD5(GUID)[:N] на длинах 8/10/12.
    Спека выбрала 10 по выборке из 47 GUID — здесь проверяем на всём массиве.
    """
    out: dict[int, list[tuple[str, str, str]]] = {}
    for n in QR_LENGTHS:
        seen: dict[str, str] = {}
        collisions: list[tuple[str, str, str]] = []
        for uid in sorted(uids):
            code = hashlib.md5(uid.encode()).hexdigest()[:n]
            if code in seen and seen[code] != uid:
                collisions.append((code, seen[code], uid))
            else:
                seen[code] = uid
        out[n] = collisions
    return out


# --- H2: номер заказа --------------------------------------------------------

PATTERNS = [
    (re.compile(r"^\d+\s+\S"), "число + текст  (6564 Спецторг ООО)"),
    (re.compile(r"^\d+$"), "только число  (6564)"),
    (re.compile(r"^[А-Яа-яA-Za-z]+\d*[-_]\d+"), "ПРЕФИКС + число  (ЛД00-006564)"),
    (re.compile(r"^\D+\d+"), "текст + число"),
    (re.compile(r"^\s*$"), "пусто"),
]


def classify_order_raw(raw: str) -> str:
    for rx, label in PATTERNS:
        if rx.match(raw or ""):
            return label
    return "прочее"


# --- Сбор массива ------------------------------------------------------------

def collect(root: Path, report: DatasetReport) -> list[Detail]:
    """Распарсить все .xbir рекурсивно."""
    all_details: list[Detail] = []
    files = sorted(root.rglob("*.xbir")) + sorted(root.rglob("*.XBIR"))

    for path in files:
        report.files_total += 1
        try:
            details, vr = parse_xbir(path)
        except Exception as e:                       # noqa: BLE001
            report.files_failed.append(f"{path.name}: {type(e).__name__}: {e}")
            continue

        report.rows_total += vr.rows_total
        report.rows_parsed += vr.rows_parsed
        report.row_errors.extend(f"{path.name}: {e}" for e in vr.errors[:20])
        all_details.extend(details)

    return all_details


def analyse(details: list[Detail], report: DatasetReport,
            plan_operations: list[str] | None = None) -> None:
    per_order: Counter = Counter()
    uids: set[str] = set()

    for d in details:
        uids.add(d.detail_uid)
        report.order_raw_patterns[classify_order_raw(d.order_raw)] += 1
        if d.order_num is None:
            report.order_num_failures.append(d.order_raw)
        else:
            per_order[d.order_num] += 1

        report.qty_distribution[d.qty] += 1
        report.thickness_distribution[d.thickness] += 1
        for e in (d.edge_l1, d.edge_l2, d.edge_w1, d.edge_w2):
            if e > 0:
                report.edge_distribution[e] += 1
        if d.drill_through + d.drill_blind + d.drill_end > 0:
            report.drill_details += 1

    report.uid_total = len(uids)
    report.orders = dict(per_order)
    report.qr_collisions = check_qr_collisions(uids)

    # H3: покрытие правил
    if plan_operations:
        for op in plan_operations:
            report.operations_in_plan[op] += 1
            if find_rule(op) is None:
                report.operations_without_rule.append(op)
        report.operations_without_rule = sorted(set(report.operations_without_rule))

    # Правила, под которые в массиве нет ни одной детали
    for rule in OPERATION_RULES:
        if rule.stage in ("quality", "warehouse", "packing", "delivery"):
            continue
        if not any(rule.detail_predicate(d) for d in details):
            report.rules_without_details.append(rule.operation_1c)

    # Детали, не попадающие ни под одно подетальное правило
    per_detail_rules = [r for r in OPERATION_RULES
                        if r.stage in ("cutting", "edging", "drilling")]
    for d in details:
        if not any(r.detail_predicate(d) for r in per_detail_rules):
            report.details_no_operation += 1


# --- Отчёт -------------------------------------------------------------------

def format_report(r: DatasetReport) -> str:
    L: list[str] = []
    add = L.append

    add("=" * 68)
    add("ВАЛИДАЦИЯ МАССИВА ДАННЫХ — проверка гипотез спецификаций")
    add("=" * 68)
    add("")
    add(f"Файлов .xbir:      {r.files_total}  (сбой разбора: {len(r.files_failed)})")
    add(f"Строк всего:       {r.rows_total}")
    add(f"Строк распарсено:  {r.rows_parsed}")
    add(f"Уникальных GUID:   {r.uid_total}")
    add(f"Заказов:           {len(r.orders)}")
    add("")

    # H1
    add("-" * 68)
    add("H1. Уникальность QR = MD5(GUID)[:N]")
    add("-" * 68)
    for n in QR_LENGTHS:
        col = r.qr_collisions.get(n, [])
        mark = "OK " if not col else "СБОЙ"
        add(f"  [{mark}] длина {n:>2}: коллизий {len(col)}")
        for c in col[:5]:
            add(f"          {c[0]}: {c[1]} ↔ {c[2]}")
    chosen = r.qr_collisions.get(10, [])
    add("")
    add(f"  ВЫВОД: длина 10 (принята в спеке) — "
        f"{'подтверждена' if not chosen else 'НЕ ГОДИТСЯ, нужно удлинять'}")
    add("")

    # H2
    add("-" * 68)
    add("H2. Извлечение номера заказа (корневой риск проекта)")
    add("-" * 68)
    for pattern, cnt in r.order_raw_patterns.most_common():
        flag = " ← РИСК" if "ПРЕФИКС" in pattern else ""
        add(f"  {cnt:>7}  {pattern}{flag}")
    add("")
    add(f"  Не извлечён номер: {len(r.order_num_failures)}")
    for raw in r.order_num_failures[:5]:
        add(f"      • '{raw}'")
    risky = sum(c for p, c in r.order_raw_patterns.items() if "ПРЕФИКС" in p)
    add("")
    if risky:
        add(f"  ВЫВОД: {risky} строк с префиксом — regex \\d+ возьмёт ПЕРВОЕ число")
        add("         и вернёт неверный заказ. Нормализатор обязателен (Этап B).")
    else:
        add("  ВЫВОД: префиксов не встретилось, текущая регулярка достаточна.")
    add("")

    # H3
    add("-" * 68)
    add("H3. Покрытие операций правилами")
    add("-" * 68)
    if r.operations_in_plan:
        add(f"  Операций в плане 1С: {len(r.operations_in_plan)}")
        add(f"  Без правила:         {len(r.operations_without_rule)}")
        for op in r.operations_without_rule[:20]:
            add(f"      • {op}")
    else:
        add("  План 1С не передан (--plan) — проверка пропущена.")
    add("")
    if r.rules_without_details:
        add("  Правила без единой подходящей детали в массиве:")
        for op in r.rules_without_details:
            add(f"      • {op}")
    add("")
    add(f"  Деталей вне всех подетальных правил: {r.details_no_operation}")
    add("")

    # H4
    add("-" * 68)
    add("H4. Распределение qty (режим сканирования)")
    add("-" * 68)
    total = sum(r.qty_distribution.values()) or 1
    for qty, cnt in sorted(r.qty_distribution.items())[:12]:
        add(f"  qty={qty:<4} {cnt:>7}  ({100 * cnt / total:.1f}%)")
    scans = sum(q * c for q, c in r.qty_distribution.items())
    add("")
    add(f"  Всего физических деталей (Σ qty): {scans}")
    add(f"  Сканов на заказ в среднем: "
        f"{scans / max(len(r.orders), 1):.0f}")
    add("")

    # H5
    add("-" * 68)
    add("H5. Пригодность данных")
    add("-" * 68)
    err_pct = 100 * (r.rows_total - r.rows_parsed) / max(r.rows_total, 1)
    add(f"  Битых строк: {r.rows_total - r.rows_parsed} ({err_pct:.2f}%)")
    for e in r.row_errors[:10]:
        add(f"      • {e}")
    add("")
    add(f"  Толщины ДСП: " + ", ".join(
        f"{t}мм×{c}" for t, c in sorted(r.thickness_distribution.items())))
    add(f"  Толщины кромки: " + ", ".join(
        f"{e}×{c}" for e, c in sorted(r.edge_distribution.items())))
    add(f"  Деталей со сверлением: {r.drill_details}")
    add("")

    return "\n".join(L)


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(
        description="Валидация гипотез спецификаций на массиве .xbir")
    p.add_argument("root", help="папка с данными (рекурсивно ищет .xbir)")
    p.add_argument("--out", default="out", help="папка для отчётов")
    p.add_argument("--plan", default=None,
                   help="xlsx плана 1С (для проверки покрытия операций)")
    args = p.parse_args(argv)

    root = Path(args.root)
    if not root.is_dir():
        print(f"Ошибка: папка не найдена: {root}", file=sys.stderr)
        return 1

    report = DatasetReport()
    print(f"Сканирую {root} …")
    details = collect(root, report)

    if not details:
        print("Не найдено ни одной детали. Проверьте, что в папке есть .xbir.",
              file=sys.stderr)
        return 1

    plan_ops = None
    if args.plan:
        plan_ops = _read_plan_operations(Path(args.plan))

    analyse(details, report, plan_ops)
    text = format_report(report)
    print()
    print(text)

    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)
    (out / "dataset_report.txt").write_text(text, encoding="utf-8")
    (out / "dataset_report.json").write_text(
        json.dumps(_jsonable(report), ensure_ascii=False, indent=2),
        encoding="utf-8")

    print(f"→ {out / 'dataset_report.txt'}")
    print(f"→ {out / 'dataset_report.json'}")
    return 0


def _read_plan_operations(path: Path) -> list[str]:
    """Достать список операций из выгрузки плана 1С (колонка operation_name)."""
    try:
        import openpyxl
    except ImportError:
        print("openpyxl не установлен — план 1С пропущен", file=sys.stderr)
        return []
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    for candidate in ("operation_name", "операция", "номенклатура"):
        if candidate in headers:
            idx = headers.index(candidate)
            return [str(row[idx].value).strip()
                    for row in ws.iter_rows(min_row=2)
                    if row[idx].value]
    print(f"В {path.name} не найдена колонка operation_name", file=sys.stderr)
    return []


def _jsonable(r: DatasetReport) -> dict:
    d = dict(r.__dict__)
    for k, v in d.items():
        if isinstance(v, Counter):
            d[k] = {str(kk): vv for kk, vv in v.most_common()}
        elif isinstance(v, dict) and k == "qr_collisions":
            d[k] = {str(kk): vv for kk, vv in v.items()}
    return d


if __name__ == "__main__":
    sys.exit(main())
