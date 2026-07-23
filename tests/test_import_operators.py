"""
Тесты импорта операторов из выгрузки 1С (src/import_operators.py).

Проверяют генерацию стабильных ID, отсев мусора и корректную
транслитерацию на реальных примерах имён из выгрузки цеха.
"""
from __future__ import annotations

import pytest

from import_operators import (
    looks_like_name,
    make_operator_id,
    transliterate,
)


# --- Транслитерация ----------------------------------------------------------

@pytest.mark.parametrize("ru,expected_start", [
    ("Мазеин", "mazein"),
    ("Жданов", "zhdanov"),
    ("Шляпникова", "shlyapnikova"),
    ("Кольцов", "kolcov"),
])
def test_transliterate(ru, expected_start):
    assert transliterate(ru).strip().startswith(expected_start)


# --- Генерация ID ------------------------------------------------------------

def test_operator_id_format():
    taken = set()
    assert make_operator_id("Мазеин Петр Иванович", taken) == "op_mazein_p_i"
    assert make_operator_id("Малышев Андрей Юрьевич", taken) == "op_malyshev_a_y"


def test_operator_id_is_stable():
    """Один и тот же человек → один и тот же ID при повторных запусках."""
    assert make_operator_id("Ивашкин Николай Алексеевич", set()) == \
           make_operator_id("Ивашкин Николай Алексеевич", set())


def test_operator_id_collision_gets_suffix():
    """Два однофамильца с одинаковыми инициалами не получат один ID."""
    taken = set()
    first = make_operator_id("Малышев Андрей Юрьевич", taken)
    second = make_operator_id("Малышев Алексей Яковлевич", taken)
    assert first != second
    assert first == "op_malyshev_a_y"
    assert second == "op_malyshev_a_y_2"


def test_brigade_gets_readable_id():
    oid = make_operator_id("Бригада (Саня Сергей)", set())
    assert oid.startswith("op_brigada")


# --- Отсев мусора ------------------------------------------------------------

@pytest.mark.parametrize("value", [
    "Жданов Андрей Александрович",
    "Бригада (Саня Сергей)",
    "Жданова Татьяна",
])
def test_real_names_pass(value):
    assert looks_like_name(value)


@pytest.mark.parametrize("value", [
    "asv",          # служебное значение из выгрузки
    "",
    "  ",
    "12",
])
def test_junk_rejected(value):
    assert not looks_like_name(value)


# --- Импорт в БД -------------------------------------------------------------

def test_import_writes_operators(tmp_path):
    """Импорт из синтетической выгрузки заносит операторов в справочник."""
    import openpyxl
    from import_operators import import_operators
    from storage import Storage

    xlsx = tmp_path / "plan.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Номер заказа", "Операция цеха", "Исполнитель"])
    ws.append(["ЛД00-001", "Облицовывание кромки 19/0,8", "Мазеин Петр Иванович"])
    ws.append(["ЛД00-001", "Облицовывание кромки 19/0,8", "Мазеин Петр Иванович"])
    ws.append(["ЛД00-002", "Раскрой плиты 16мм", "Жданов Андрей Александрович"])
    ws.append(["ЛД00-003", "Операция ОТК", "asv"])
    wb.save(xlsx)

    storage = Storage(tmp_path / "op.db")
    report = import_operators(storage, xlsx)

    ids = {oid for oid, _ in report.imported}
    assert "op_mazein_p_i" in ids
    assert "op_zhdanov_a_a" in ids
    assert ("asv", 1) in report.review          # мусор на ревью, не заведён

    assert storage.get_operator("op_mazein_p_i")["full_name"] == "Мазеин Петр Иванович"
    assert len(storage.get_operators()) == 2
    storage.close()


def test_area_filter(tmp_path):
    """С фильтром по участку заводятся только операторы этого участка."""
    import openpyxl
    from import_operators import import_operators
    from storage import Storage

    plan = tmp_path / "plan.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Номер заказа", "Операция цеха", "Исполнитель"])
    ws.append(["ЛД00-001", "Облицовывание кромки 19/0,8", "Кромлёнщик Иван Иванович"])
    ws.append(["ЛД00-002", "Раскрой плиты 16мм", "Раскройщик Пётр Петрович"])
    wb.save(plan)

    areas = tmp_path / "areas.xlsx"
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Операции по участкам"
    ws2.append(["Операция цеха", "Участок"])
    ws2.append(["Облицовывание кромки 19/0,8", "Кромление"])
    ws2.append(["Раскрой плиты 16мм", "Раскрой"])
    wb2.save(areas)

    storage = Storage(tmp_path / "op.db")
    report = import_operators(storage, plan, area="Кромление", areas_path=areas)

    names = {name for _, name in report.imported}
    assert "Кромлёнщик Иван Иванович" in names
    assert "Раскройщик Пётр Петрович" not in names   # другой участок
    storage.close()
